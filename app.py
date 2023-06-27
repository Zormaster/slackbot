import os, re
from slack_bolt import App
from slack_bolt.adapter.socket_mode import SocketModeHandler
from dotenv import load_dotenv
from openpyxl import load_workbook
import pandas as pd

load_dotenv()
SLACK_WORKSPACE = os.getenv("SLACK_WORKSPACE")
SLACK_BOT_TOKEN = os.getenv("SLACK_BOT_TOKEN")
SLACK_APP_TOKEN = os.getenv("SLACK_APP_TOKEN")
SLACK_USER_TOKEN = os.getenv("SLACK_USER_TOKEN")
EXCEL_FILE="Slack Access List.xlsx"

DEBUG=False
#DEBUG=True

# Initializes your app with your bot token and socket mode handler
print(SLACK_WORKSPACE)
print(SLACK_WORKSPACE)
app = App(token=os.environ.get("SLACK_BOT_TOKEN"))

def send_invites(filename, sheet):
    df = pd.read_excel(filename, sheet_name=sheet)
    # Set channels to be added to invitation
    all_channels = app.client.conversations_list(
        limit=500
    )
    join_channels = ''
    for i in all_channels['channels']:
        if i['name'] == 'general':
            join_channels += f"{i['id']},"
        if i['name'] == 'random':
            join_channels += f"{i['id']},"
        # if sheet name found in channel name, add channel to list
        if re.match(f'.*{sheet.lower()}.*', i['name']):
            join_channels += f"{i['id']},"
    
    # Get team ID (required for admin_users_invite function)
    team_id = app.client.team_info()['team']['id']

    for i in df.index:
        if not pd.isna(df['Email'][i]):
            print(f"{df['First Name'][i]} {df['Last Name'][i]}")
            print(df['Email'][i])
            print(join_channels)
            print(team_id)
            # # This will send an invitation and should be uncommented when the app is ready to use
            # app.client.admin_users_invite(
            #     channel_ids=join_channels,
            #     email=df['Email'][i],
            #     team_id=team_id,
            #     real_name=f"{df['First Name']} {df['Last Name']}",
            #     resend=True
            # )

def set_display_name(id, email, fullname, displayname, manual):
    if manual:
        displayname = ' '.join(displayname)
    print('User Update: id=', id, ', email=', email, ', fn=', fullname, ', dn=', displayname)    
    userprofile = app.client.users_profile_set(
            token = SLACK_USER_TOKEN,
            user = id,
            profile = {'real_name':fullname, 'display_name': displayname}
        )
    if DEBUG:
        print(userprofile)

@app.command("/invite_bulk")
def invite_bulk(ack, command, respond):
    ack()

    # Load excel sheet
    filename = EXCEL_FILE
    wb = load_workbook(filename)
    sheets = re.split("[,\s]", command['text'])
    
    # Initialize message
    msg = ''

    # If command parameters exist, send invites to all 
    if sheets[0] == '':
        for i, sheet in enumerate(wb.sheetnames[:-1]):
            send_invites(filename, sheet)
            msg += f"Invitations sent to {sheet}\n"

    # Else send invites only to specified groups
    else:
        for i in sheets:
            if i in wb.sheetnames:
                send_invites(filename, i)
                msg += f"Invitations sent to {i}\n"

            else:
                msg += f"{i} not found in sheets.\n"

    # Post message to slack channel
    respond(msg)

@app.command("/set_display_name")
def handle_set_display_name(ack, client, command, respond):
    ack()
 
    #Parse parameters embedded in command['text']
    params = re.split(",\s", command['text'])
    msg = ''
    print('/set_display_name called with params: ', params)

    # Get user profiles
    # If command parameters exist, update single profile 
    if params[0] != '':
        email, fullname, displayname = params[0], params[1], params[2:]
        userprofile = client.users_lookupByEmail(
            email = email
        )
        set_display_name(userprofile['user']['id'], email, fullname, displayname, True)
        msg += f"User {email} updated with {fullname} and {displayname}\n"
    
    # Else bulk update from Excel sheet
    else:
        # Load excel sheet
        filename = EXCEL_FILE
        respond(f"Loading file {filename}\n")
        print(f"Loading file {filename}")
        wb = load_workbook(filename)
        
        # Initialize message
        for i, sheet in enumerate(wb.sheetnames):
            # send_invites(filename, sheet)
            print(f"---- Reading sheet: {sheet} ----")
            msg += f"---- Reading sheet: {sheet} ----\n"
            df = pd.read_excel(filename, sheet_name=sheet)
            for i in df.index:
                if not pd.isna(df['Email'][i]):
                    email = f"{df['Email'][i]}"
                    fullname = f"{df['Full Name'][i]}"
                    displayname = f"{df['Display Name'][i]}"
                    if fullname == 'nan': 
                        fullname=""
                    if displayname == 'nan':
                        displayname=""
                    try:
                        userprofile = client.users_lookupByEmail(
                            email = email
                        )
                        set_display_name(userprofile['user']['id'], email, fullname, displayname, False)
                        msg += f"User {email} updated with {fullname} and {displayname}\n"
                        # print(f"{df['First Name'][i]} {df['Last Name'][i]}")
                        # print(df['Email'][i])
                    except:
                        print("Exception")
                        msg += f"User {df['Email'][i]} update failed\n"
                    
            msg += f"---- Updated Sheet: {sheet} ----\n"

    respond(msg)

# Start your app
if __name__ == "__main__":
    SocketModeHandler(app, os.environ["SLACK_APP_TOKEN"]).start()
